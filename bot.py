#!/usr/bin/env python3
"""
КосыТату — Бот инвентаризации студии по плетению косичек
@KosaDH_bot
"""

import os
import sqlite3
import logging
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ConversationHandler, filters, ContextTypes
)
from telegram.constants import ParseMode

logging.basicConfig(format="%(asctime)s | %(levelname)s | %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

TOKEN   = os.environ.get("BOT_TOKEN", "REPLACE_WITH_YOUR_TOKEN")
DB_PATH = os.environ.get("DB_PATH", "inventory.db")

CATEGORIES = [
    "Зизи", "Гофре", "Изи Брейд", "Канекалон",
    "Комплект", "Украшения", "Заколки и ободки", "Татуировки",
]

UNIT_MAP = {
    "Зизи":             "нитей",
    "Гофре":            "нитей",
    "Изи Брейд":        "локонов",
    "Канекалон":        "пачек",
    "Комплект":         "нитей",
    "Украшения":        "штук",
    "Заколки и ободки": "штук",
    "Татуировки":       "штук",
}

(
    ST_MAIN,
    ST_INV_CAT, ST_INV_ITEM, ST_INV_QTY,
    ST_WO_CAT, ST_WO_ITEM, ST_WO_QTY,
    ST_ADD_CAT, ST_ADD_NAME, ST_ADD_PHOTO,
    ST_MANAGE,
    ST_REPORT,
) = range(12)

# ══════════════════════════════════════════════════════════════════════════════
#  БАЗА ДАННЫХ
# ══════════════════════════════════════════════════════════════════════════════

def _conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c

def init_db():
    with _conn() as db:
        db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            tg_id    INTEGER PRIMARY KEY,
            username TEXT,
            name     TEXT,
            role     TEXT NOT NULL DEFAULT 'pending',
            joined   TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS items (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            category  TEXT NOT NULL,
            name      TEXT NOT NULL,
            photo_id  TEXT,
            unit      TEXT NOT NULL,
            added_at  TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS sessions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL,
            type        TEXT NOT NULL DEFAULT 'inventory',
            started_at  TEXT NOT NULL,
            finished_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(tg_id)
        );
        CREATE TABLE IF NOT EXISTS entries (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            item_id    INTEGER NOT NULL,
            quantity   REAL    NOT NULL,
            noted_at   TEXT    NOT NULL,
            FOREIGN KEY(session_id) REFERENCES sessions(id),
            FOREIGN KEY(item_id)    REFERENCES items(id)
        );
        """)

def now():   return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
def today(): return datetime.now().strftime("%d.%m.%Y %H:%M")
def qstr(q): return str(int(q)) if q == int(q) else str(q)

def db_user(tg_id):
    with _conn() as db:
        return db.execute("SELECT * FROM users WHERE tg_id=?", (tg_id,)).fetchone()

def db_count_admins():
    with _conn() as db:
        return db.execute("SELECT COUNT(*) FROM users WHERE role='admin'").fetchone()[0]

def db_upsert_user(tg_id, username, name, role="pending"):
    with _conn() as db:
        db.execute("""
            INSERT INTO users(tg_id,username,name,role,joined)
            VALUES(?,?,?,?,?)
            ON CONFLICT(tg_id) DO UPDATE
            SET username=excluded.username, name=excluded.name
        """, (tg_id, username or "", name, role, now()))

def db_set_role(tg_id, role):
    with _conn() as db:
        db.execute("UPDATE users SET role=? WHERE tg_id=?", (role, tg_id))

def db_admins():
    with _conn() as db:
        return db.execute("SELECT * FROM users WHERE role IN ('admin','director')").fetchall()

def db_pending():
    with _conn() as db:
        return db.execute("SELECT * FROM users WHERE role='pending'").fetchall()

def db_items(category):
    with _conn() as db:
        return db.execute("SELECT * FROM items WHERE category=? ORDER BY name", (category,)).fetchall()

def db_item(item_id):
    with _conn() as db:
        return db.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()

def db_add_item(category, name, photo_id, unit):
    with _conn() as db:
        db.execute("INSERT INTO items(category,name,photo_id,unit,added_at) VALUES(?,?,?,?,?)",
                   (category, name, photo_id, unit, now()))

def db_new_session(user_id, stype="inventory"):
    with _conn() as db:
        c = db.execute("INSERT INTO sessions(user_id,type,started_at) VALUES(?,?,?)",
                       (user_id, stype, now()))
        return c.lastrowid

def db_close_session(sid):
    with _conn() as db:
        db.execute("UPDATE sessions SET finished_at=? WHERE id=?", (now(), sid))

def db_add_entry(sid, item_id, qty):
    with _conn() as db:
        db.execute("INSERT INTO entries(session_id,item_id,quantity,noted_at) VALUES(?,?,?,?)",
                   (sid, item_id, qty, now()))

def db_session_entries(sid):
    with _conn() as db:
        return db.execute("""
            SELECT e.quantity, i.name, i.category, i.unit
            FROM entries e JOIN items i ON i.id=e.item_id
            WHERE e.session_id=?
        """, (sid,)).fetchall()

def db_latest_inventory():
    """Последний остаток по каждой позиции из инвентаризаций."""
    with _conn() as db:
        return db.execute("""
            SELECT i.category, i.name, i.unit, e.quantity, s.finished_at, u.name AS user_name
            FROM entries e
            JOIN items i    ON i.id=e.item_id
            JOIN sessions s ON s.id=e.session_id
            JOIN users u    ON u.tg_id=s.user_id
            WHERE s.finished_at IS NOT NULL AND s.type='inventory'
              AND e.id = (
                SELECT MAX(e2.id) FROM entries e2
                JOIN sessions s2 ON s2.id=e2.session_id
                WHERE e2.item_id=e.item_id AND s2.finished_at IS NOT NULL AND s2.type='inventory'
              )
            ORDER BY i.category, i.name
        """).fetchall()

def db_writeoffs_today():
    """Списания за сегодня."""
    today_date = datetime.now().strftime("%Y-%m-%d")
    with _conn() as db:
        return db.execute("""
            SELECT i.category, i.name, i.unit, SUM(e.quantity) as total, u.name AS user_name
            FROM entries e
            JOIN items i    ON i.id=e.item_id
            JOIN sessions s ON s.id=e.session_id
            JOIN users u    ON u.tg_id=s.user_id
            WHERE s.type='writeoff' AND s.finished_at IS NOT NULL
              AND date(s.finished_at) = ?
            GROUP BY e.item_id
            ORDER BY i.category, i.name
        """, (today_date,)).fetchall()

def db_last_two_sessions(stype="inventory"):
    with _conn() as db:
        return db.execute("""
            SELECT s.id, s.finished_at, u.name AS user_name
            FROM sessions s JOIN users u ON u.tg_id=s.user_id
            WHERE s.finished_at IS NOT NULL AND s.type=?
            ORDER BY s.finished_at DESC LIMIT 2
        """, (stype,)).fetchall()

def db_sessions_for_excel():
    with _conn() as db:
        return db.execute("""
            SELECT s.id, s.finished_at, s.type, u.name AS user_name
            FROM sessions s JOIN users u ON u.tg_id=s.user_id
            WHERE s.finished_at IS NOT NULL
            ORDER BY s.finished_at DESC LIMIT 50
        """).fetchall()

# ══════════════════════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ══════════════════════════════════════════════════════════════════════════════

def is_staff(u): return u and u["role"] in ("admin", "director", "employee")
def is_mgmt(u):  return u and u["role"] in ("admin", "director")

ROLE_LABEL = {
    "admin":    "👑 Администратор",
    "director": "🏢 Директор",
    "employee": "👤 Сотрудник",
    "pending":  "⏳ Ожидает одобрения",
}

def main_kb(role):
    btns = [
        [InlineKeyboardButton("📦 Инвентаризация", callback_data="inv_start")],
        [InlineKeyboardButton("✂️ Списание",        callback_data="wo_start")],
    ]
    if role in ("admin", "director"):
        btns += [
            [InlineKeyboardButton("➕ Добавить товар",       callback_data="add_item")],
            [InlineKeyboardButton("👥 Управление доступом",  callback_data="manage_users")],
        ]
    btns.append([InlineKeyboardButton("📊 Отчёты", callback_data="report")])
    return InlineKeyboardMarkup(btns)

def cat_kb(prefix):
    rows = []
    for i in range(0, len(CATEGORIES), 2):
        row = [InlineKeyboardButton(CATEGORIES[i], callback_data=f"{prefix}:{CATEGORIES[i]}")]
        if i + 1 < len(CATEGORIES):
            row.append(InlineKeyboardButton(CATEGORIES[i+1], callback_data=f"{prefix}:{CATEGORIES[i+1]}"))
        rows.append(row)
    rows.append([InlineKeyboardButton("◀️ Главное меню", callback_data="back_main")])
    return InlineKeyboardMarkup(rows)

def items_kb(category, entries, finish_cb="inv_finish", back_cb="back_cats"):
    items = db_items(category)
    btns = []
    for it in items:
        label = it["name"]
        if it["id"] in entries:
            label += f"  ✅ {entries[it['id']]}"
        btns.append([InlineKeyboardButton(label, callback_data=f"item:{it['id']}")])
    btns.append([InlineKeyboardButton("◀️ К категориям", callback_data=back_cb)])
    btns.append([InlineKeyboardButton("✅ Завершить", callback_data=finish_cb)])
    return InlineKeyboardMarkup(btns)

def report_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Текущие остатки",       callback_data="rep_current")],
        [InlineKeyboardButton("✂️ Списания за сегодня",   callback_data="rep_writeoffs")],
        [InlineKeyboardButton("📉 Разница инвентаризаций", callback_data="rep_diff")],
        [InlineKeyboardButton("📥 Скачать Excel",          callback_data="rep_excel")],
        [InlineKeyboardButton("◀️ Главное меню",           callback_data="back_main")],
    ])

def back_report_kb():
    return InlineKeyboardMarkup([[InlineKeyboardButton("◀️ К отчётам", callback_data="report")]])

# ══════════════════════════════════════════════════════════════════════════════
#  ХЭНДЛЕРЫ
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_start(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    u = upd.effective_user
    row = db_user(u.id)
    if not row:
        role = "admin" if db_count_admins() == 0 else "pending"
        db_upsert_user(u.id, u.username, u.full_name, role)
        row = db_user(u.id)
        if role == "admin":
            await upd.message.reply_text(
                "👑 Добро пожаловать, Администратор!\nВы первый пользователь.",
                reply_markup=main_kb("admin")
            )
            return ST_MAIN
        for admin in db_admins():
            try:
                await ctx.bot.send_message(
                    admin["tg_id"],
                    f"🆕 *Запрос доступа*\n👤 {u.full_name}\n📱 @{u.username or '—'}\n🆔 `{u.id}`",
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=InlineKeyboardMarkup([
                        [InlineKeyboardButton("✅ Сотрудник", callback_data=f"apr:employee:{u.id}"),
                         InlineKeyboardButton("🏢 Директор", callback_data=f"apr:director:{u.id}")],
                        [InlineKeyboardButton("❌ Отклонить", callback_data=f"apr:rejected:{u.id}")]
                    ])
                )
            except Exception: pass
        await upd.message.reply_text("👋 Запрос отправлен администратору. Ожидайте.")
        return ConversationHandler.END
    role = row["role"]
    if role == "pending":
        await upd.message.reply_text("⏳ Запрос ещё не одобрён.")
        return ConversationHandler.END
    if role == "rejected":
        await upd.message.reply_text("❌ В доступе отказано.")
        return ConversationHandler.END
    await upd.message.reply_text(
        f"Привет, {u.first_name}! {ROLE_LABEL.get(role, role)}\n\nЧто будем делать?",
        reply_markup=main_kb(role)
    )
    return ST_MAIN

async def cb_approve(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query
    approver = db_user(q.from_user.id)
    if not is_mgmt(approver):
        await q.answer("❌ Нет прав", show_alert=True); return
    parts = q.data.split(":")
    new_role, tid = parts[1], int(parts[2])
    db_set_role(tid, new_role)
    await q.answer()
    labels = {"employee": "Сотрудник ✅", "director": "Директор 🏢", "rejected": "Отклонён ❌"}
    await q.edit_message_text(f"Готово! Роль: {labels.get(new_role, new_role)}")
    msg = "❌ В доступе отказано." if new_role == "rejected" else f"✅ Доступ одобрен! Роль: {labels[new_role]}\nНапишите /start"
    try: await ctx.bot.send_message(tid, msg)
    except Exception: pass

async def cb_main(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query
    u = db_user(q.from_user.id)
    await q.answer()

    if q.data == "back_main":
        if not is_staff(u): return ConversationHandler.END
        await q.edit_message_text("Главное меню:", reply_markup=main_kb(u["role"]))
        return ST_MAIN

    if q.data == "inv_start":
        if not is_staff(u): return ST_MAIN
        sid = db_new_session(q.from_user.id, "inventory")
        ctx.user_data.update({"sid": sid, "inv": {}, "cat": None, "mode": "inv"})
        await q.edit_message_text("📦 Инвентаризация\nВыберите категорию:", reply_markup=cat_kb("inv_cat"))
        return ST_INV_CAT

    if q.data == "wo_start":
        if not is_staff(u): return ST_MAIN
        sid = db_new_session(q.from_user.id, "writeoff")
        ctx.user_data.update({"sid": sid, "inv": {}, "cat": None, "mode": "wo"})
        await q.edit_message_text("✂️ Списание\nВыберите категорию:", reply_markup=cat_kb("wo_cat"))
        return ST_WO_CAT

    if q.data == "add_item":
        if not is_mgmt(u): return ST_MAIN
        await q.edit_message_text("➕ Добавить товар\nВыберите категорию:", reply_markup=cat_kb("add_cat"))
        return ST_ADD_CAT

    if q.data == "manage_users":
        if not is_mgmt(u): return ST_MAIN
        return await show_manage(q, ctx)

    if q.data == "report":
        await q.edit_message_text("📊 Выберите отчёт:", reply_markup=report_kb())
        return ST_REPORT

    return ST_MAIN

# ══════════════════════════════════════════════════════════════════════════════
#  ИНВЕНТАРИЗАЦИЯ
# ══════════════════════════════════════════════════════════════════════════════

async def cb_inv_cat(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query; await q.answer()
    if q.data == "back_main":
        u = db_user(q.from_user.id)
        await q.edit_message_text("Главное меню:", reply_markup=main_kb(u["role"]))
        return ST_MAIN
    if q.data in ("back_cats", "inv_finish"):
        if q.data == "inv_finish":
            return await do_finish(q, ctx, "inventory")
        await q.edit_message_text("📦 Выберите категорию:", reply_markup=cat_kb("inv_cat"))
        return ST_INV_CAT
    cat = q.data.split(":", 1)[1]
    ctx.user_data["cat"] = cat
    items = db_items(cat)
    if not items:
        await q.edit_message_text(f"📂 *{cat}*\n\n⚠️ Нет товаров.", parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("◀️ К категориям", callback_data="back_cats")],
                [InlineKeyboardButton("✅ Завершить", callback_data="inv_finish")]
            ]))
        return ST_INV_CAT
    await q.edit_message_text(f"📂 *{cat}*\nВыберите позицию:", parse_mode=ParseMode.MARKDOWN,
        reply_markup=items_kb(cat, ctx.user_data.get("inv", {}), "inv_finish", "back_cats"))
    return ST_INV_ITEM

async def cb_inv_item(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query; await q.answer()
    if q.data == "back_cats":
        await q.edit_message_text("📦 Выберите категорию:", reply_markup=cat_kb("inv_cat"))
        return ST_INV_CAT
    if q.data == "inv_finish":
        return await do_finish(q, ctx, "inventory")
    item_id = int(q.data.split(":")[1])
    item = db_item(item_id)
    ctx.user_data["item_id"] = item_id
    caption = f"📌 *{item['name']}*\n📂 {item['category']}\n📏 {item['unit']}\n\nВведите текущий остаток:"
    back_btn = InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="back_to_cat")]])
    if item["photo_id"]:
        try: await q.message.delete()
        except Exception: pass
        sent = await ctx.bot.send_photo(q.from_user.id, item["photo_id"], caption=caption,
                                         parse_mode=ParseMode.MARKDOWN, reply_markup=back_btn)
        ctx.user_data["photo_msg_id"] = sent.message_id
    else:
        await q.edit_message_text(caption, parse_mode=ParseMode.MARKDOWN, reply_markup=back_btn)
    return ST_INV_QTY

async def cb_back_to_cat(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query; await q.answer()
    cat = ctx.user_data.get("cat", "")
    mode = ctx.user_data.get("mode", "inv")
    photo_mid = ctx.user_data.pop("photo_msg_id", None)
    if photo_mid:
        try: await ctx.bot.delete_message(q.from_user.id, photo_mid)
        except Exception: pass
    if mode == "wo":
        await ctx.bot.send_message(q.from_user.id, f"✂️ *{cat}*\nВыберите позицию:",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=items_kb(cat, ctx.user_data.get("inv", {}), "wo_finish", "wo_back_cats"))
        return ST_WO_ITEM
    await ctx.bot.send_message(q.from_user.id, f"📂 *{cat}*\nВыберите позицию:",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=items_kb(cat, ctx.user_data.get("inv", {}), "inv_finish", "back_cats"))
    return ST_INV_ITEM

async def msg_qty(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = upd.message.text.replace(",", ".").strip()
    try:
        qty = float(text)
        if qty < 0: raise ValueError
    except ValueError:
        await upd.message.reply_text("⚠️ Введите число (например: 42 или 0.5)")
        return ST_INV_QTY
    item_id = ctx.user_data["item_id"]
    item = db_item(item_id)
    sid = ctx.user_data["sid"]
    db_add_entry(sid, item_id, qty)
    ctx.user_data["inv"][item_id] = f"{qstr(qty)} {item['unit']}"
    photo_mid = ctx.user_data.pop("photo_msg_id", None)
    if photo_mid:
        try: await ctx.bot.delete_message(upd.effective_user.id, photo_mid)
        except Exception: pass
    cat = ctx.user_data["cat"]
    await upd.message.reply_text(f"✅ *{item['name']}*: {qstr(qty)} {item['unit']}", parse_mode=ParseMode.MARKDOWN)
    await upd.message.reply_text(f"📂 *{cat}*:", parse_mode=ParseMode.MARKDOWN,
        reply_markup=items_kb(cat, ctx.user_data["inv"], "inv_finish", "back_cats"))
    return ST_INV_ITEM

# ══════════════════════════════════════════════════════════════════════════════
#  СПИСАНИЕ
# ══════════════════════════════════════════════════════════════════════════════

async def cb_wo_cat(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query; await q.answer()
    if q.data == "back_main":
        u = db_user(q.from_user.id)
        await q.edit_message_text("Главное меню:", reply_markup=main_kb(u["role"]))
        return ST_MAIN
    if q.data in ("wo_back_cats", "wo_finish"):
        if q.data == "wo_finish":
            return await do_finish(q, ctx, "writeoff")
        await q.edit_message_text("✂️ Списание\nВыберите категорию:", reply_markup=cat_kb("wo_cat"))
        return ST_WO_CAT
    cat = q.data.split(":", 1)[1]
    ctx.user_data["cat"] = cat
    items = db_items(cat)
    if not items:
        await q.edit_message_text(f"📂 *{cat}*\n\n⚠️ Нет товаров.", parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("◀️ К категориям", callback_data="wo_back_cats")],
                [InlineKeyboardButton("✅ Завершить", callback_data="wo_finish")]
            ]))
        return ST_WO_CAT
    await q.edit_message_text(f"✂️ *{cat}*\nСколько израсходовано?", parse_mode=ParseMode.MARKDOWN,
        reply_markup=items_kb(cat, ctx.user_data.get("inv", {}), "wo_finish", "wo_back_cats"))
    return ST_WO_ITEM

async def cb_wo_item(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query; await q.answer()
    if q.data == "wo_back_cats":
        await q.edit_message_text("✂️ Списание\nВыберите категорию:", reply_markup=cat_kb("wo_cat"))
        return ST_WO_CAT
    if q.data == "wo_finish":
        return await do_finish(q, ctx, "writeoff")
    item_id = int(q.data.split(":")[1])
    item = db_item(item_id)
    ctx.user_data["item_id"] = item_id
    caption = f"✂️ *{item['name']}*\n📂 {item['category']}\n📏 {item['unit']}\n\nВведите количество израсходованного:"
    back_btn = InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="back_to_cat")]])
    if item["photo_id"]:
        try: await q.message.delete()
        except Exception: pass
        sent = await ctx.bot.send_photo(q.from_user.id, item["photo_id"], caption=caption,
                                         parse_mode=ParseMode.MARKDOWN, reply_markup=back_btn)
        ctx.user_data["photo_msg_id"] = sent.message_id
    else:
        await q.edit_message_text(caption, parse_mode=ParseMode.MARKDOWN, reply_markup=back_btn)
    return ST_WO_QTY

async def msg_wo_qty(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = upd.message.text.replace(",", ".").strip()
    try:
        qty = float(text)
        if qty < 0: raise ValueError
    except ValueError:
        await upd.message.reply_text("⚠️ Введите число (например: 5 или 0.5)")
        return ST_WO_QTY
    item_id = ctx.user_data["item_id"]
    item = db_item(item_id)
    sid = ctx.user_data["sid"]
    db_add_entry(sid, item_id, qty)
    ctx.user_data["inv"][item_id] = f"{qstr(qty)} {item['unit']}"
    photo_mid = ctx.user_data.pop("photo_msg_id", None)
    if photo_mid:
        try: await ctx.bot.delete_message(upd.effective_user.id, photo_mid)
        except Exception: pass
    cat = ctx.user_data["cat"]
    await upd.message.reply_text(f"✂️ Списано: *{item['name']}*: {qstr(qty)} {item['unit']}", parse_mode=ParseMode.MARKDOWN)
    await upd.message.reply_text(f"✂️ *{cat}*:", parse_mode=ParseMode.MARKDOWN,
        reply_markup=items_kb(cat, ctx.user_data["inv"], "wo_finish", "wo_back_cats"))
    return ST_WO_ITEM

# ══════════════════════════════════════════════════════════════════════════════
#  ЗАВЕРШЕНИЕ СЕССИИ
# ══════════════════════════════════════════════════════════════════════════════

async def do_finish(q, ctx: ContextTypes.DEFAULT_TYPE, stype: str):
    sid = ctx.user_data.get("sid")
    if sid: db_close_session(sid)
    inv = ctx.user_data.get("inv", {})
    u = db_user(q.from_user.id)
    emoji = "📦" if stype == "inventory" else "✂️"
    label = "Инвентаризация" if stype == "inventory" else "Списание"
    cat_data: dict[str, list] = {}
    with _conn() as db:
        for item_id, qty_label in inv.items():
            it = db.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()
            if it:
                cat_data.setdefault(it["category"], []).append(f"• {it['name']}: {qty_label}")
    lines = [f"{emoji} *{label} завершено*", f"👤 {u['name']}", f"🕐 {today()}\n"]
    if cat_data:
        for cat, rows in cat_data.items():
            lines.append(f"*{cat}:*")
            lines.extend(rows)
            lines.append("")
    else:
        lines.append("_(нет данных)_")
    summary = "\n".join(lines)
    await q.message.reply_text(summary, parse_mode=ParseMode.MARKDOWN)
    for admin in db_admins():
        if admin["tg_id"] != q.from_user.id:
            try:
                await ctx.bot.send_message(admin["tg_id"], f"🔔 *{label}*\n\n{summary}", parse_mode=ParseMode.MARKDOWN)
            except Exception: pass
    db_u = db_user(q.from_user.id)
    await q.message.reply_text("Главное меню:", reply_markup=main_kb(db_u["role"]))
    ctx.user_data.clear()
    return ST_MAIN

# ══════════════════════════════════════════════════════════════════════════════
#  ДОБАВЛЕНИЕ ТОВАРА
# ══════════════════════════════════════════════════════════════════════════════

async def cb_add_cat(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query; await q.answer()
    if q.data == "back_main":
        u = db_user(q.from_user.id)
        await q.edit_message_text("Главное меню:", reply_markup=main_kb(u["role"]))
        return ST_MAIN
    cat = q.data.split(":", 1)[1]
    ctx.user_data["new_cat"] = cat
    ctx.user_data["new_unit"] = UNIT_MAP[cat]
    await q.edit_message_text(f"📂 *{cat}*\n\nВведите название (цвет) товара:", parse_mode=ParseMode.MARKDOWN)
    return ST_ADD_NAME

async def msg_add_name(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["new_name"] = upd.message.text.strip()
    await upd.message.reply_text("📸 Отправьте фото или пропустите:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⏭ Без фото", callback_data="skip_photo")]]))
    return ST_ADD_PHOTO

async def msg_add_photo(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["new_photo"] = upd.message.photo[-1].file_id
    return await _save_item(upd, ctx)

async def cb_skip_photo(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query; await q.answer()
    ctx.user_data["new_photo"] = None
    return await _save_item(upd, ctx)

async def _save_item(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    cat = ctx.user_data["new_cat"]; name = ctx.user_data["new_name"]
    photo = ctx.user_data.get("new_photo"); unit = ctx.user_data["new_unit"]
    db_add_item(cat, name, photo, unit)
    u = db_user(upd.effective_user.id)
    msg = upd.callback_query.message if upd.callback_query else upd.message
    await msg.reply_text(f"✅ *Добавлено!*\n📂 {cat} / 🏷 {name} / 📏 {unit}",
        parse_mode=ParseMode.MARKDOWN, reply_markup=main_kb(u["role"]))
    ctx.user_data.clear()
    return ST_MAIN

# ══════════════════════════════════════════════════════════════════════════════
#  УПРАВЛЕНИЕ ПОЛЬЗОВАТЕЛЯМИ
# ══════════════════════════════════════════════════════════════════════════════

async def show_manage(q, ctx):
    pending = db_pending()
    if pending:
        text = f"👥 *Ожидают одобрения: {len(pending)}*"
        btns = []
        for p in pending:
            uname = f"@{p['username']}" if p["username"] else "нет username"
            text += f"\n👤 {p['name']} ({uname})"
            btns += [
                [InlineKeyboardButton(f"✅ Сотрудник — {p['name'][:18]}", callback_data=f"apr:employee:{p['tg_id']}")],
                [InlineKeyboardButton("🏢 Директор", callback_data=f"apr:director:{p['tg_id']}"),
                 InlineKeyboardButton("❌ Отклонить", callback_data=f"apr:rejected:{p['tg_id']}")],
            ]
    else:
        text = "👥 Нет заявок."; btns = []
    btns.append([InlineKeyboardButton("◀️ Главное меню", callback_data="back_main")])
    await q.edit_message_text(text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(btns))
    return ST_MANAGE

async def cb_manage(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query; await q.answer()
    if q.data == "back_main":
        u = db_user(q.from_user.id)
        await q.edit_message_text("Главное меню:", reply_markup=main_kb(u["role"]))
        return ST_MAIN
    if q.data.startswith("apr:"):
        await cb_approve(upd, ctx)
        pending = db_pending()
        if pending: return await show_manage(q, ctx)
        u = db_user(q.from_user.id)
        await ctx.bot.send_message(q.from_user.id, "👥 Больше нет заявок.", reply_markup=main_kb(u["role"]))
        return ST_MAIN
    return ST_MANAGE

# ══════════════════════════════════════════════════════════════════════════════
#  ОТЧЁТЫ
# ══════════════════════════════════════════════════════════════════════════════

async def cb_report(upd: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = upd.callback_query; await q.answer()

    if q.data == "back_main":
        u = db_user(q.from_user.id)
        await q.edit_message_text("Главное меню:", reply_markup=main_kb(u["role"]))
        return ST_MAIN

    if q.data == "report":
        await q.edit_message_text("📊 Выберите отчёт:", reply_markup=report_kb())
        return ST_REPORT

    if q.data == "rep_current":
        rows = db_latest_inventory()
        writeoffs = {r["name"]: r for r in db_writeoffs_today()}
        if not rows:
            await q.edit_message_text("📋 Данных пока нет.", reply_markup=back_report_kb())
            return ST_REPORT
        lines = ["📋 *Текущие остатки*\n"]
        cur_cat = None
        for r in rows:
            if r["category"] != cur_cat:
                cur_cat = r["category"]
                lines.append(f"\n*{cur_cat}:*")
            qty = qstr(r["quantity"])
            wo = writeoffs.get(r["name"])
            if wo:
                remain = r["quantity"] - wo["total"]
                lines.append(f"  • {r['name']}: {qty} {r['unit']} _(списано: {qstr(wo['total'])}, остаток: {qstr(remain)})_")
            else:
                lines.append(f"  • {r['name']}: {qty} {r['unit']}")
        await q.edit_message_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN, reply_markup=back_report_kb())
        return ST_REPORT

    if q.data == "rep_writeoffs":
        rows = db_writeoffs_today()
        if not rows:
            await q.edit_message_text("✂️ Сегодня списаний не было.", reply_markup=back_report_kb())
            return ST_REPORT
        lines = [f"✂️ *Списания за {datetime.now().strftime('%d.%m.%Y')}*\n"]
        cur_cat = None
        for r in rows:
            if r["category"] != cur_cat:
                cur_cat = r["category"]
                lines.append(f"\n*{cur_cat}:*")
            lines.append(f"  • {r['name']}: {qstr(r['total'])} {r['unit']}")
        await q.edit_message_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN, reply_markup=back_report_kb())
        return ST_REPORT

    if q.data == "rep_diff":
        sessions = db_last_two_sessions("inventory")
        if len(sessions) < 2:
            await q.edit_message_text("⚠️ Нужно минимум 2 инвентаризации.", reply_markup=back_report_kb())
            return ST_REPORT
        newer, older = sessions[0], sessions[1]
        ne = {e["name"]: e for e in db_session_entries(newer["id"])}
        oe = {e["name"]: e for e in db_session_entries(older["id"])}
        all_names = sorted(set(ne) | set(oe))
        lines = [
            "📉 *Разница по остаткам*\n",
            f"Старая: {older['user_name']} ({older['finished_at'][:10]})",
            f"Новая:  {newer['user_name']} ({newer['finished_at'][:10]})\n",
        ]
        changed = False
        for name in all_names:
            nq = ne[name]["quantity"] if name in ne else 0.0
            oq = oe[name]["quantity"] if name in oe else 0.0
            diff = nq - oq
            if diff == 0: continue
            changed = True
            unit = (ne.get(name) or oe.get(name))["unit"]
            sign = "+" if diff > 0 else ""
            emoji = "🔺" if diff > 0 else "🔻"
            lines.append(f"{emoji} {name}: {sign}{qstr(diff)} {unit}")
        if not changed: lines.append("✅ Изменений нет")
        await q.edit_message_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN, reply_markup=back_report_kb())
        return ST_REPORT

    if q.data == "rep_excel":
        sessions = db_sessions_for_excel()
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "КосыТату"
        hfont = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="1B4F72")
        halign = Alignment(horizontal="center")
        for col, h in enumerate(["Дата","Тип","Сотрудник","Категория","Наименование","Кол-во","Единица"], 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hfont; cell.fill = hfill; cell.alignment = halign
        row_n = 2
        for s in sessions:
            stype_label = "Инвентаризация" if s["type"] == "inventory" else "Списание"
            for e in db_session_entries(s["id"]):
                ws.cell(row=row_n, column=1, value=s["finished_at"][:16])
                ws.cell(row=row_n, column=2, value=stype_label)
                ws.cell(row=row_n, column=3, value=s["user_name"])
                ws.cell(row=row_n, column=4, value=e["category"])
                ws.cell(row=row_n, column=5, value=e["name"])
                ws.cell(row=row_n, column=6, value=e["quantity"])
                ws.cell(row=row_n, column=7, value=e["unit"])
                row_n += 1
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 4, 45)
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"kosa_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        await ctx.bot.send_document(q.from_user.id, document=buf, filename=fname, caption="📊 Отчёт КосыТату")
        await q.edit_message_text("✅ Excel отправлен ⬆️", reply_markup=back_report_kb())
        return ST_REPORT

    return ST_REPORT

# ══════════════════════════════════════════════════════════════════════════════
#  ЗАПУСК
# ══════════════════════════════════════════════════════════════════════════════

def main():
    init_db()
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CallbackQueryHandler(cb_approve, pattern=r"^apr:"))
    conv = ConversationHandler(
        entry_points=[CommandHandler("start", cmd_start)],
        states={
            ST_MAIN: [CallbackQueryHandler(cb_main)],
            ST_INV_CAT: [CallbackQueryHandler(cb_inv_cat, pattern=r"^(inv_cat:|back_cats$|back_main$|inv_finish$)")],
            ST_INV_ITEM: [CallbackQueryHandler(cb_inv_item, pattern=r"^(item:|back_cats$|inv_finish$)")],
            ST_INV_QTY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, msg_qty),
                CallbackQueryHandler(cb_back_to_cat, pattern=r"^back_to_cat$"),
            ],
            ST_WO_CAT: [CallbackQueryHandler(cb_wo_cat, pattern=r"^(wo_cat:|wo_back_cats$|back_main$|wo_finish$)")],
            ST_WO_ITEM: [CallbackQueryHandler(cb_wo_item, pattern=r"^(item:|wo_back_cats$|wo_finish$)")],
            ST_WO_QTY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, msg_wo_qty),
                CallbackQueryHandler(cb_back_to_cat, pattern=r"^back_to_cat$"),
            ],
            ST_ADD_CAT: [CallbackQueryHandler(cb_add_cat, pattern=r"^(add_cat:|back_main$)")],
            ST_ADD_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, msg_add_name)],
            ST_ADD_PHOTO: [
                MessageHandler(filters.PHOTO, msg_add_photo),
                CallbackQueryHandler(cb_skip_photo, pattern=r"^skip_photo$"),
            ],
            ST_MANAGE: [CallbackQueryHandler(cb_manage)],
            ST_REPORT: [CallbackQueryHandler(cb_report)],
        },
        fallbacks=[CommandHandler("start", cmd_start)],
        per_user=True, per_chat=True, allow_reentry=True,
    )
    app.add_handler(conv)
    logger.info("Бот запущен!")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
